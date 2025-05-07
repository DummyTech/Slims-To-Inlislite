import csv
from openpyxl import load_workbook, Workbook
import copy
from pathlib import Path
import os

### ORIGINAL CODE BY "DummyTech (Farrell)"

### FILE NAMES
# Books
books_file_reference = "Books Format.xlsx" #Fixed const
books_file_format = "Books Format (Template).xlsx" #Fixed const
biblioFile = "senayan_biblio_export.csv"
itemFile = "senayan_item_export.csv"

# Members
members_file_reference = "Members Format.xlsx" #Fixed const
members_file_format = "Members Format (Template).xlsx" #Fixed const
membersFile = "senayan_member_export.csv"

# Output folder name
outputFolder = "Results"



### PATH
local_dir = script_path = Path(__file__).resolve().parent

# Books
books_path_reference = Path(f"{local_dir}\\Template\\{books_file_reference}")
books_path_format = Path(f"{local_dir}\\Template\\{books_file_format}")

biblio_path = Path(f"{local_dir}\\{biblioFile}")
print(biblio_path.exists())
item_path = Path(f"{local_dir}\\{itemFile}")

# Members
members_path_reference = Path(f"{local_dir}\\Template\\{members_file_reference}")
members_path_format = Path(f"{local_dir}\\Template\\{members_file_format}")
members_path = Path(f"{local_dir}\\{membersFile}")

# Output Path 
output_dir = Path(f"{local_dir}\\{outputFolder}")
os.makedirs(output_dir, exist_ok=True)

def ask_for_path():
    global biblio_path, biblio_path, itemFile, item_path, membersFile, members_path

    print("### ASKING FILE NAME")
    print("Note = File should be in adjacent or in one directory with the script")
    while not biblio_path.exists():
        biblioFile = input("(Include the .csv at the end)\nEnter your biblio export name? ")
        biblio_path = Path(f"{local_dir}\\{biblioFile}")

    while not item_path.exists():
        itemFile = input("(Include the .csv at the end)\nEnter your item export name? ")
        item_path = Path(f"{local_dir}\\{itemFile}")

    while not members_path.exists():
        membersFile = input("(Include the .csv at the end)\nEnter your members export name? ")
        members_path = Path(f"{local_dir}\\{membersFile}")

def run_books():
    splitter = 0
    fileIndex = 1
    wb_ref = load_workbook(books_path_reference)
    ws_ref = wb_ref.active 
    wb = load_workbook(books_path_format)
    ws = wb.active

    bookNumber = 0
    indexNumber = 0


    with open(biblio_path, mode='r', newline='', encoding='utf-8') as file2, open(item_path, mode='r', newline='', encoding='utf-8') as file1:
        print("Opening csv")
        items = csv.reader(file1, delimiter='^')
        biblio = csv.reader(file2, delimiter='^')


        for row in items:
            if splitter >= 100:
                print(f"Saving Output({fileIndex}).xlsx")
                splitter = 0
                wb.save(f'{output_dir}\\output({fileIndex}).xlsx')
                fileIndex += 1
                indexNumber = 0
                wb = load_workbook(books_path_format)
                ws = wb.active

            bookNumber += 1
            print(f"Processing book = \"{row[18]}\"")
            file2.seek(0)
            indexNumber += 1
            splitter += 1
            BiblioRow = None
            for i in biblio:
                if i[0] == row[18]:
                    BiblioRow = i
            #print(BiblioRow)

            #print(row)
            new_row = [''] * 42
            new_row[0] = indexNumber # Nomor index

            cache = row[4].split("-")
            day = cache[2]
            month = cache[1]
            year = cache[0]
            new_row[1] = f"{day}-{month}-{year}" # TGL Pengadaan

            new_row[2] = row[0] # No. Induk
            new_row[3] = row[0] # No. Barcode
            new_row[4] = row[0] # No. RFID
            new_row[5] = "Pembelian" # Jenis sumber
            new_row[6] = "---Belum ditentukan---" # Nama sumber
            new_row[7] = "IDR" # Mata uang
            new_row[8] = row[13] # Harga
            new_row[9] = row[7] # Kode lokasi perpustakaan
            new_row[10] = "0000" # Kode lokasi ruangan
            new_row[11] = "Dapat dipinjam" # Akses
            new_row[12] = "Koleksi Umum" # Kategori
            new_row[13] = "Buku" # Media
            new_row[14] = "Tersedia" # Ketersediaan
            new_row[15] = row[1] # Nomor panggil eksemplar
            new_row[16] = "Monograf" # Jenis bahan
            new_row[17] = row[18] # Judul utama
            new_row[18] = "" # Anak Judul
            new_row[19] = BiblioRow[14] # Pernyataan Tanggung jawab
            new_row[20] =  BiblioRow[15][1:-1].replace("><", ", ")# Tajuk Pengarang
            new_row[21] = "" # Tajung Pengarang badan koperasi
            new_row[22] = "" # Pengarang tambahan nama orang
            new_row[23] = "" # Pengarang tambahan nama badan
            new_row[24] = BiblioRow[2] # Edisi
            new_row[25] = BiblioRow[10] # Kota terbit
            new_row[26] = BiblioRow[4] # Penerbit
            new_row[27] = BiblioRow[5] # Tahun terbit
            new_row[28] = BiblioRow[6] # Jumlah halaman
            new_row[29] = BiblioRow[6] # Dimensi
            new_row[30] = BiblioRow[3] # ISBN
            new_row[31] = "" # ISSN
            new_row[32] = "" # ISMN
            new_row[33] = BiblioRow[11] # No Doc
            new_row[34] = BiblioRow[8] # Nomor panggil katalog
            new_row[35] = BiblioRow[12] # Abstrak

            if BiblioRow[9] == "Indonesia": # Bahasa
                new_row[36] = "ind"
            elif BiblioRow[9] == "English":
                new_row[36] = "eng"
            elif BiblioRow[9] == "":
                new_row[36] = ""
            else:
                raise Exception("Language is either english or indonesia")

            new_row[37] = BiblioRow[16][1:-1].replace("><", ", ") # Subjek Topik
            new_row[38] = BiblioRow[7] # Edisi serial
            new_row[39] = "" # TGL Terbit Edisi Serial
            new_row[40] = "" # Bahan Sertaan (Serial)
            new_row[41] = "" # Keterangan Lain (Serial)

            ws.append(new_row)

            for col_index, source_cell in enumerate(ws_ref[3], start=1):
                dest_cell = ws.cell(row=indexNumber+2, column = col_index)
                dest_cell.font = copy.copy(source_cell.font)
                dest_cell.border = copy.copy(source_cell.border)
                dest_cell.fill = copy.copy(source_cell.fill)
                dest_cell.number_format = source_cell.number_format
                dest_cell.protection = copy.copy(source_cell.protection)
                dest_cell.alignment = copy.copy(source_cell.alignment)


    wb.save(f'{output_dir}\\output({fileIndex}).xlsx')
    print(f"All book formated ({bookNumber} Books, ({fileIndex}) File Output)")

def run_members():
    wb = load_workbook(members_path_format)
    ws = wb.active
    wb_ref = load_workbook(members_path_reference)
    ws_ref = wb_ref.active
    indexNumber = 0

    with open(members_path, mode='r', newline='', encoding='utf-8') as file1:
        print("Opening csv")
        keanggotaan = csv.reader(file1, delimiter='^')

        for row in keanggotaan:
            print(f"Processing member = \"{row[1]}\"")
            indexNumber += 1

            new_row = [''] * 49
            new_row[0] = indexNumber # Nomor index
            new_row[1] = row[0] # No anggota
            new_row[2] = row[1] # Nama
            
            # Tanggal lahir (Required)
            if row[16] == "":
                new_row[4] = "00-00-0000"
            else:
                cache = row[16].split("-")
                new_row[4] = f"{cache[2]}-{cache[1]}-{cache[0]}"

            new_row[5] = row[5] # Alamat
            new_row[19] = row[11] # No HP
            new_row[21] = "Kartu Pelajar" # Jenis identitas (Required)

            # No Identitas (Required)
            if row[10] == "":
                new_row[22] = f"NO_ID_TIDAK_ADA{indexNumber}"
            else:
                new_row[22] = row[10] 

            # Jenis kelamin (Required)
            if row[2] == "1":
                new_row[23] = "Laki-laki"
            elif row[2] == "0":
                new_row[23] = "Perempuan"
            else:
                raise Exception("Gender error")
            
            new_row[27] = row[4] # Alamat email
            new_row[28] = "Pelajar" # Jenis Anggota (Required)
            new_row[29] = row[7] # Pendidikan terakhir

            # Tanggal Pendaftaran (Required)
            if row[14] == "":
                new_row[31] = "00-00-0000"
            else:
                cache = row[14].split("-")
                new_row[31] = f"{cache[2]}-{cache[1]}-{cache[0]}"

            # Tanggal Akhir Berlaku (Required)
            if row[15] == "":
                new_row[32] = "00-00-0000"
            else:
                cache = row[15].split("-")
                new_row[32] = f"{cache[2]}-{cache[1]}-{cache[0]}"

            new_row[33] = "Baru" # Jenis Permohonan 
            new_row[34] = "Aktif" # Status Anggota (Required)
            new_row[47] = row[3] # Kelas Siswa
            new_row[48] = row[9] # Photo URL

            ws.append(new_row)

            for col_index, source_cell in enumerate(ws_ref[2], start=1):
                dest_cell = ws.cell(row=indexNumber+1, column=col_index)
                dest_cell.font = copy.copy(source_cell.font)
                dest_cell.border = copy.copy(source_cell.border)
                dest_cell.fill = copy.copy(source_cell.fill)
                dest_cell.number_format = source_cell.number_format
                dest_cell.protection = copy.copy(source_cell.protection)
                dest_cell.alignment = copy.copy(source_cell.alignment)
                #dest_cell.value = copy.copy(source_cell.value)

        
    wb.save(f'{output_dir}\\output(Keangotaan).xlsx')
    print(f"All members formated ({indexNumber} members)")



run = True
if not books_path_reference.exists():
    print("Books format file is missing.\nPlease name sample file to \"Books Format.xlsx\" and\ncheck if file is the template folder.\n")
    run = False
if not books_path_format.exists():
    print("Books template file is missing.\nPlease name sample file to \"Books Format (Template).xlsx\" and\ncheck if file is the template folder.\n")
    run = False
if not members_path_reference.exists():
    print("Members format file is missing.\nPlease name sample file to \"Members Format.xlsx\" and\ncheck if file is the template folder.\n")
    run = False
if not members_path_format.exists():
    print("Members template file is missing.\nPlease name sample file to \"Members Format (Template).xlsx\" and\ncheck if file is the template folder.\n")
    run = False


while run:
    userInput = input(" - Rewrite biblio & item (1)\n - Rewrite member (2)\n - Do both (3)\nYour input? ") 

    if userInput == "1" or userInput == "2" or userInput == "3":
        break

if run:
    try:
        if userInput == "1":
            if not biblio_path.exists() or not item_path.exists():
                ask_for_path()

            run_books()

        elif userInput == "2":
            if not members_path.exists():
                ask_for_path()

            run_members()

        elif userInput == "3":
            if not biblio_path.exists() or not item_path.exists() or not members_path.exists():                     
                ask_for_path()

            run_books()
            run_members()

    except Exception as e:
        print("An error occurred:", e)
        run = False

    finally:
        if run:
            print("\n\nNOTE: You need to \"resave\" or \"save-as\" the file using a table software like WPS or Microsoft Excel. Since the data output is unclean.\nFolder named \"Re Saved\" created for storing the resaved files.\nThanks for using our software :)\nDummyTech >u<\n")
            os.makedirs("Re Saved", exist_ok=True)

input("Press Enter to exit...")         
